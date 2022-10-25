from django.shortcuts import render, redirect
from django.contrib import messages
from mainapp.forms import DataForm
from mainapp import models
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import AuthenticationForm
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib.auth import login, authenticate
from .forms import SignupForm
from django.contrib.sites.shortcuts import get_current_site
from django.utils.encoding import force_bytes, force_str # force_text
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.template.loader import render_to_string
from .tokens import account_activation_token
from django.contrib.auth.models import User
from django.contrib.auth import get_user_model
from django.core.mail import EmailMessage


def finalactivation(request, uidb64, token):
    User = get_user_model()
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except(TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None
    if user is not None and account_activation_token.check_token(user, token):
        user.is_active = True
        user.save()

        messages.success(request, 'Thank you for your email confirmation. Now you can login your account.')
        return redirect ('mainapp:login')
#            return HttpResponse('Thank you for your email confirmation. Now you can login your account.')
    else:
        messages.error(request, 'Activation link is invalid!')
        return redirect ('mainapp:login')

    return redirect ('mainapp:login')



def signup(request):

    if request.method == 'POST':
        form = SignupForm(request.POST)
        email = request.POST['email']
        if User.objects.filter(email=email).exists():
            messages.error(request, 'That email is already in use')
        elif form.is_valid():
            user = form.save(commit=False)
            user.is_active = False
            user.save()
            # to get the domain of the current site
            current_site = get_current_site(request)
            mail_subject = 'Activation link has been sent to your email id'
            message = render_to_string('mainapp/acc_active_email.html', {
                'user': user,
                'domain': current_site.domain,
                'uid':urlsafe_base64_encode(force_bytes(user.pk)),
                'token':account_activation_token.make_token(user),

            })
            to_email = form.cleaned_data.get('email')
            email = EmailMessage(
                mail_subject, message, to=[to_email]
            )
            email.send()
            messages.success(request, 'Please confirm your email address to complete the registration'
            )
            return render (request,'mainapp/welcome.html')
#            return HttpResponse('Please confirm your email address to complete the registration')
    else:
        form = SignupForm()
    return render(request, 'mainapp/register.html', {'form': form})


def login_request(request):

    if request.method == 'POST':
        login_form = AuthenticationForm(request, data=request.POST)
        if login_form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            
            user = authenticate(username=username, password=password,
                                backend='django.contrib.auth.backends.ModelBackend')
            if user is not None:
                login(request, user)
  
                return redirect('mainapp:home')
            else:
                messages.error(request, 'Wrong Password or Username')
        else:
            messages.error(request, 'Wrong Password or Username')
    login_form = AuthenticationForm()
    return render(request, template_name='mainapp/login.html', context={'login_form': login_form})



#@login_required(login_url='mainapp:login')
def home(request):
    form = DataForm()
    queryset = models.Data.objects.all()
    if request.method == 'POST':
        form = DataForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('home')
    context = {'title': 'Simple App', 'form': form, 'posts': queryset}
    return render(request, 'mainapp/index.html', context=context)


from openpyxl import Workbook
from django.http import HttpResponse
from openpyxl.styles import Font


#@login_required(login_url='mainapp:login')
def export_data(request):
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Data.xlsx"'

    # create workbook
    wb = Workbook()

    sheet = wb.active

    # stylize header row
    # 'id','title', 'quantity','pub_date'

    c1 = sheet.cell(row = 1, column = 1)
    c1.value = "id"
    c1.font = Font(bold=True)

    c2 = sheet.cell(row= 1 , column = 2)
    c2.value = "title"
    c2.font = Font(bold=True)

    c3 = sheet.cell(row= 1 , column = 3)
    c3.value = "quantity"
    c3.font = Font(bold=True)

    c4 = sheet.cell(row= 1 , column = 4)
    c4.value = "pub_date"
    c4.font = Font(bold=True)

    # export data to Excel
    rows = models.Data.objects.all().values_list('id','category', 'quantity','pub_date',)
    for row_num, row in enumerate(rows, 1):
        # row is just a tuple
        for col_num, value in enumerate(row):
            c5 = sheet.cell(row=row_num+1, column=col_num+1)
            c5.value = value

    wb.save(response)

    return response

